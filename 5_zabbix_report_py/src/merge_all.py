import csv
import re
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from modules.utils import *

# =======================
# FILES
# =======================
# PERIOD_PERIOD_DAYS = 28
TREND_FILE = f"reports/zbx_trends.csv"
DISK_UTIL_FILE = f"reports/zbx_disks.csv"
FS_FILE = "reports/zbx_disks_fs.csv"
CPU_SPIKES_FILE = f"reports/zbx_cpu_spikes.csv"
OUT_FILE = f"reports/merged_all_{PERIOD_DAYS}d.xlsx"

# =======================
# HIGHLIGHT TOGGLE
# =======================
HIGHLIGHT = True # ← can be turned off

# =======================
# FUNCTIONS
# =======================

def find_hostid_field(fieldnames):
    for f in fieldnames:
        clean = f.strip().lstrip("\ufeff")
        if clean.lower() == "hostid":
            return f
    raise KeyError("HostID column not found!")

def insert_columns(fields, inserts):
    new_fields = fields.copy()
    for item in inserts:
        name = item["name"]
        if name in new_fields:
            new_fields.remove(name)
        if item.get("after"):
            col = item["after"]
            idx = new_fields.index(col) + 1
            new_fields.insert(idx, name)
        else:
            new_fields.append(name)
    return new_fields

def convert_text_to_number(cell):
    """Text-number -> Excel number (Converts a comma-decimal string to float)"""
    v = cell.value
    if v is None:
        return

    if isinstance(v, str):
        v_strip = v.strip()

        if re.match(r"^-?\d+,\d+$", v_strip):
            try:
                # Replace comma with dot for Python to convert to float
                cell.value = float(v_strip.replace(",", "."))
            except:
                pass

        elif re.match(r"^-?\d+$", v_strip):
            try:
                cell.value = int(v_strip)
            except:
                pass


# ============================
# 🔥 HIGHLIGHTING
# ============================
def apply_highlighting(ws, fieldnames):

    yellow = PatternFill("solid", fgColor="FFF2CC")

    try:
        idx_ram = fieldnames.index("%_RAM_Util")
        idx_ram_max = fieldnames.index("%_RAM_Util_MAX")
        idx_spike_max = fieldnames.index("CPU_Spike_Max_s")
        idx_spike_total = fieldnames.index("CPU_Spikes_Total_s")
        idx_spikes = fieldnames.index("CPU_Spikes_Count")
        idx_disk_used = fieldnames.index("%_Disk_Used_Agg")
    except ValueError as e:
        print(f"⚠️ Warning: Highlighting skipped because field {e} was not found in final columns.")
        return


    for row in ws.iter_rows(min_row=2):

        # %RAM_Util / MAX > 80
        for cell in (row[idx_ram], row[idx_ram_max]):
            try:
                val = str(cell.value).replace(",", ".")
                if float(val) > 80:
                    cell.fill = yellow
            except:
                pass

        # CPU spikes fields > 80
        for cell in (row[idx_spike_max], row[idx_spike_total]):
            try:
                val = str(cell.value).replace(",", ".")
                if float(val) > 80:
                    cell.fill = yellow
            except:
                pass

        try:
            if int(str(row[idx_spikes].value)) > 0:
                row[idx_spikes].fill = yellow
        except:
            pass

        # Disk_used > 80
        try:
            val = str(row[idx_disk_used].value).replace(",", ".")
            if float(val) > 80:
                row[idx_disk_used].fill = yellow
        except:
            pass


# =======================
# 1) BACKEND + LOGICAL DISKS
# =======================
backend_counts = Counter()
logical_sets = {}

with open(DISK_UTIL_FILE, encoding="utf-8", newline="") as f:
    reader = csv.DictReader(f, delimiter=';')
    disk_hostid = find_hostid_field(reader.fieldnames)

    for row in reader:
        host_id = row[disk_hostid].strip()
        if not host_id:
            continue

        backend_counts[host_id] += 1

        disk_raw = row["Disk"].strip()
        logical = disk_raw.split()[-1] if ":" in disk_raw else disk_raw
        logical_sets.setdefault(host_id, set()).add(logical)

# =======================
# 2) FILESYSTEM COUNT AND DISK TOTALS AGGREGATION
# =======================
fs_counts = Counter()
disk_totals = {} 

with open(FS_FILE, encoding="utf-8", newline="") as f:
    reader = csv.DictReader(f, delimiter=';')
    fs_hostid = find_hostid_field(reader.fieldnames)

    for row in reader:
        host_id = row[fs_hostid].strip()
        if not host_id:
            continue

        # 2.1 FS Count
        fs_counts[host_id] += 1
        
        # 2.2 Disk Totals Aggregation 
        try:
            # Convert comma-decimal string to dot for float conversion in Python
            total_gb_str = row.get("Total_GB", "0").replace(",", ".")
            free_gb_str = row.get("Free_GB", "0").replace(",", ".")
            
            total_gb = float(total_gb_str)
            free_gb = float(free_gb_str)

        except (ValueError, KeyError, AttributeError):
            continue

        disk_totals.setdefault(host_id, {'Total_GB': 0.0, 'Free_GB': 0.0})
        disk_totals[host_id]['Total_GB'] += total_gb
        disk_totals[host_id]['Free_GB'] += free_gb

# =======================
# 3) CPU SPIKES
# =======================
cpu_spikes = {}

with open(CPU_SPIKES_FILE, encoding="utf-8", newline="") as f:
    reader = csv.DictReader(f, delimiter=';')
    spike_hostid = find_hostid_field(reader.fieldnames)

    for row in reader:
        host_id = row[spike_hostid].strip()
        if not host_id:
            continue

        cpu_spikes[host_id] = {
            "CPU_Spikes_Count": row["CPU_Spikes_Count"].strip(),
            "CPU_Spike_Max_s": row["CPU_Spike_Max_s"].strip(),
            "CPU_Spikes_Total_s": row["CPU_Spikes_Total_s"].strip(),
        }


# =======================
# 4) READ TRENDS + BUILD XLSX
# =======================
with open(TREND_FILE, encoding="utf-8", newline="") as fin:

    reader = csv.DictReader(fin, delimiter=';')
    trend_hostid = find_hostid_field(reader.fieldnames)

    inserts = [ 
        {"name": "Disk_Count_Backend", "after": "%_RAM_Util_MAX"},
        {"name": "Disk_Count_Logical", "after": "Disk_Count_Backend"},
        {"name": "FS_Count", "after": "Disk_Count_Logical"},
        
        # AGGREGATED FIELDS (UPDATED)
        {"name": "Disk_Total_Agg_GB", "after": "FS_Count"},
        {"name": "Disk_Used_Agg_GB", "after": "Disk_Total_Agg_GB"}, 
        # {"name": "Disk_Free_Agg_GB", "after": "Disk_Used_Agg_GB"}, # COMMENTED OUT
        {"name": "%_Disk_Used_Agg", "after": "Disk_Used_Agg_GB"}, # Updated "after"
        
        {"name": "CPU_Spikes_Count", "after": "%_Disk_Used_Agg"}, 
        {"name": "CPU_Spike_Max_s", "after": "CPU_Spikes_Count"},
        {"name": "CPU_Spikes_Total_s", "after": "CPU_Spike_Max_s"},
    ]

    fieldnames = insert_columns(reader.fieldnames, inserts)

    wb = Workbook()
    ws = wb.active
    ws.title = "Zabbix Report"

    # header row
    ws.append(fieldnames)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
        cell.alignment = Alignment(horizontal="center")

    # data rows
    for row in reader:
        host_id = row[trend_hostid].strip()

        row["Disk_Count_Backend"] = backend_counts.get(host_id, "") or ""
        row["Disk_Count_Logical"] = len(logical_sets.get(host_id, [])) or ""
        row["FS_Count"] = fs_counts.get(host_id, "") or ""

        # INTEGRATION OF AGGREGATED DATA (UPDATED)
        totals = disk_totals.get(host_id, {})
        if totals:
            total_float = totals.get('Total_GB', 0.0)
            free_float = totals.get('Free_GB', 0.0)
            
            used_float = total_float - free_float
            used_pct_float = (used_float / total_float) * 100 if total_float > 0 else 0.0

            # Format float to string with 2 decimals and replace dot with comma (for Excel)
            row["Disk_Total_Agg_GB"] = f"{total_float:.2f}".replace('.', ',')
            row["Disk_Used_Agg_GB"] = f"{used_float:.2f}".replace('.', ',') 
            # row["Disk_Free_Agg_GB"] = f"{free_float:.2f}".replace('.', ',') # Free space
            row["%_Disk_Used_Agg"] = f"{used_pct_float:.2f}".replace('.', ',') 
        else:
             row["Disk_Total_Agg_GB"] = ""
             row["Disk_Used_Agg_GB"] = ""
             # row["Disk_Free_Agg_GB"] = "" # Free space
             row["%_Disk_Used_Agg"] = ""
        # ------------------------------------

        spikes = cpu_spikes.get(host_id)
        if spikes:
            row["CPU_Spikes_Count"] = spikes["CPU_Spikes_Count"] or ""
            row["CPU_Spike_Max_s"] = spikes["CPU_Spike_Max_s"] or ""
            row["CPU_Spikes_Total_s"] = spikes["CPU_Spikes_Total_s"] or ""
        else:
            row["CPU_Spikes_Count"] = ""
            row["CPU_Spike_Max_s"] = ""
            row["CPU_Spikes_Total_s"] = ""

        ws.append([row.get(col, "") for col in fieldnames])

        # → fix 'number-as-text' (CONVERTS COMMA-DECIMAL STRINGS TO EXCEL NUMBERS)
        for cell in ws[ws.max_row]:
            convert_text_to_number(cell)

# =======================
# FORMATTING
# =======================

# auto width
for col in ws.columns:
    max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

# ===========================
# HIGHLIGHTING (if enabled)
# ===========================
if HIGHLIGHT:
    apply_highlighting(ws, fieldnames)

# =======================
# SAVE FILE
# =======================
wb.save(OUT_FILE)
print("✅ DONE →", OUT_FILE)